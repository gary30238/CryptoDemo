import os
import time
import numpy as np
import requests
import datetime
import pandas as pd
from openpyxl import load_workbook
from binance import ThreadedWebsocketManager
from binance.client import Client
import gspread
from oauth2client.service_account import ServiceAccountCredentials

def notify(s):
    headers = {
        "Authorization": "Bearer " + "line_token",
        "Content-Type": "application/x-www-form-urlencoded"
    }
 
    params = {"message": s}
 
    r = requests.post("https://notify-api.line.me/api/notify",
                      headers=headers, params=params)

def get_klines_iter(symbol, interval, start=None, end=None):
    klines = client.get_klines(symbol=symbol, interval=interval, startTime=start, endTime=end, limit=5)
    
    df2 = pd.DataFrame(klines)
    #df2.drop(df2.columns[len(df2.columns)-1], axis=1, inplace=True)
    df2.columns = ['Opentime', 'Open', 'High', 'Low', 'Close', 'Volume', 'Closetime', 'Quote asset volume', 'Number of trades','Taker by base', 'Taker buy quote', 'Ig']
    #df2['T'] = pd.to_datetime(df2['Opentime'], unit='ms') + pd.to_timedelta(8, unit='H')
    #df = pd.concat([df2, df], axis=0, keys=None)
    #df.reset_index(drop=True, inplace=True)
    return df2

def getCrypts():
    file = "binance.xlsx"
    wb = load_workbook(file)
    ws = wb.active
    allCrypts = np.array([row[0].value for row in ws.iter_rows(2)])
    wb.close()
    return allCrypts

# init
api_key = os.environ.get('binance_api')
api_secret = os.environ.get('binance_secret')
client = Client(api_key, api_secret)

allCrypts = getCrypts()
tick_interval = Client.KLINE_INTERVAL_5MINUTE
#startTime = str(int(datetime.fromisoformat('2021-06-05 08:00').timestamp()*1000))
#endTime = str(int(datetime.fromisoformat('2021-06-05 16:00').timestamp()*1000))

start300 = time.time()
s = "\n"
lineAr = []
ar1 = []
ar2 = []
count = 0
while True:
    startWhile = time.time()
    startLimit = time.time()
    for crypt in allCrypts:
        runTime = time.time() - startLimit
        if count >= 20:
            print("limit"+str(runTime))
            count = 0
            if runTime < 1:
                time.sleep(1.1 - runTime)
            startLimit = time.time()
        startTime = str(int((datetime.datetime.now() - datetime.timedelta(minutes = 5)).timestamp()*1000))
        try:
            klines = client.get_klines(symbol=crypt+'USDT', interval=tick_interval, startTime=startTime, limit=20)
            count += 1
            #df = pd.DataFrame(klines)
            #df.columns = ['Opentime', 'Open', 'High', 'Low', 'Close', 'Volume', 'Closetime', 'Quote asset volume', 'Number of trades','Taker by base', 'Taker buy quote', 'Ig']
            if klines[0][4] > klines[0][1]:
                startTime = str(int((datetime.datetime.now() - datetime.timedelta(minutes = 30)).timestamp()*1000))
                klines1 = get_klines_iter(symbol=crypt+'USDT', interval=tick_interval, start=startTime)
                count += 1
                fl = klines1["High"].astype(float).max()/float(klines[0][4])
                fn = len(klines1[klines1["Close"] < klines1["Open"]])
                if fn >= 3 and fl > 1.05:
                    df2 = pd.DataFrame([[crypt, fl, fn]], columns=['Crypt', 'Fluctuation', 'Fnum'])
                    if crypt not in ar1 :
                        ar1.append(crypt)
                        lineAr.append(crypt+"-"+str(fl)+"-"+str(fn))
                if klines1["Volume"].astype(float).sum() < float(klines[0][5]):
                    if crypt not in ar2 :
                        ar2.append(crypt)
                        notify(crypt + "成交量暴增")
        except:
            print("error"+str(count))
            print("error"+str(runTime))
            time.sleep(5)
            
        if time.time() - start300 >= 300:
            notify(s + s.join(lineAr))
            lineAr = []
            ar1 = []
            ar2 = []
            start300 = time.time()

    print("while"+str(time.time()-startWhile))
