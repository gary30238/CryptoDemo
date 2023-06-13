from openpyxl import load_workbook
import numpy as np

wb = load_workbook(filename="binance.xlsx", data_only=True)
ws = wb.active

firstRow = 2
firstCol = 1
nCols = 1
nRows = ws.max_row - 1

allCells = np.array([row[0].value for row in ws.iter_rows(2)])
