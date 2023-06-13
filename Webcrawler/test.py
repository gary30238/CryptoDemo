from openpyxl import Workbook
import openpyxl
import time
import requests

file = "binance.xlsx"
wb = openpyxl.load_workbook(file)
ws = wb.active

temp = ''
for row in ws.iter_rows(0):
    for cell in row:
        if cell.value == "ETH":
            temp = cell.value
            ws.cell(row=cell.row, column=2).value = 1
            break
    if temp != '':
        temp = ''
        break

wb.save(filename = file)


